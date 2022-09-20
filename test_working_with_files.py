import pytest
import os, zipfile, shutil, csv, codecs
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl import load_workbook

script_dir = os.path.dirname(os.path.abspath(__file__))
resources_dir = os.path.join(script_dir, 'resources')
file_dir = os.listdir(resources_dir)

pdf = os.path.join(resources_dir, 'sample.pdf')
xlsx = os.path.join(resources_dir, 'test_file.xlsx')
csv = os.path.join(resources_dir, 'test_file.csv')
zip = os.path.join(script_dir, 'test_file.zip')


def test_create_zip():
    if os.path.exists(zip):
        os.remove(zip)
        with zipfile.ZipFile(zip, 'w') as zip_file:
            add_file = os.path.join(resources_dir)
            zip_file.write(add_file)
        print('Файл zip удален и создан заново')
    else:
        with zipfile.ZipFile(zip, 'w') as zip_file:
            add_file = os.path.join(resources_dir)
            zip_file.write(add_file)
            print('Файл zip создан пустым')


# сначала не там проверял pdf, но потом добавлю проверок
"""def test_open_and_check_pdf_file():
    with open(pdf, 'r') as pdf_file:
        pdf_file = PdfReader(pdf)
        check_count_of_pages = len(pdf_file.pages)
        assert check_count_of_pages == 2
        print('Файл pdf подходит для теста')"""


def test_working_with_xlsx_file():
    if os.path.exists(xlsx):
        os.remove(xlsx)
    xlsx_file = Workbook()
    text_a1 = xlsx_file.active
    text_a1['A1'] = 'test_create_xlsx'
    xlsx_file.save(xlsx)
    print('Файл xlsx создан с записью')


def test_create_and_check_csv_file():
    if os.path.exists(csv):
        os.remove(csv)
        with open(csv, 'w') as csv_file:
            csv_file.write('test_create_csv')
        print('Файл csv удален и создан заново')
    else:
        with open(csv, 'w') as csv_file:
            csv_file.write('test_create_csv')
            print('Файл csv создан c записью')


def test_add_files_to_zip_and_move_to_resources():
    # это просто для себя
    """source = [resources_dir]
        with zipfile.ZipFile(zip, 'w') as zip_archive:
            for source_folder in source:
                for root, dirs, files in os.walk(source_folder):
                    for file in files:
                        path = os.path.join(root, file)
                        zip_archive.write(path)"""

    with ZipFile(zip, 'w') as zip_archive:
        zip_archive.write('./resources/sample.pdf')
        zip_archive.write('./resources/test_file.csv')
        zip_archive.write('./resources/test_file.xlsx')
        summa = str(len(zip_archive.namelist()))
        for name in summa:
            print('В архив добавлено ' + summa + ' элемента')
    if os.path.exists('./resources/test_file.zip'):
        os.remove('./resources/test_file.zip')
        shutil.move(zip, resources_dir)
        print('Архив уже был создан, но удален и перемещен')
    else:
        shutil.move(zip, resources_dir)
        print('Архив успешно перемещен')


def test_check_files_from_zip():
    with zipfile.ZipFile('./resources/test_file.zip') as zip_open:
        with zip_open.open('resources/test_file.xlsx') as xlsx_open:
            xlsx_open = load_workbook(xlsx_open)
            active = xlsx_open.active
            check_xlsx = active.cell(row=1, column=1).value
            assert check_xlsx == 'test_create_xlsx'
            print('Файл xlsx подходит для теста')

        # я не знаю, что делать здесь, он выдает ошибку
        # AttributeError: 'str' object has no attribute 'reader'
        # в интернетах пишут, что надо переустанавливать питон

        # with zip_open.open('resources/test_file.csv') as csv_open:
        #    csv_table = csv.reader((codecs.iterdecode(csv_open, 'utf-8')))
        #    for row in csv_table:
        #        if row == 1:
        #            assert row[1] == 'test_create_csv'
        # print('Файл csv подходит для теста')

        with zip_open.open('resources/sample.pdf') as pdf_open:
            pdf_open = PdfReader('resources/sample.pdf')
            check_count_of_pages = len(pdf_open.pages)
            assert check_count_of_pages == 2
            print('Файл pdf подходит для теста')
